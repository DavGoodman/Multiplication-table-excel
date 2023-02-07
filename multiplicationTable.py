#! pyhton3
# Creates a NxN multiplication table in excel
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
wb = openpyxl.Workbook()
sheet = wb['Sheet']

multiply_number = input()
assert multiply_number.isdecimal()

for i in range(2, int(multiply_number) + 2):
    sheet["A" + str(i)] = i - 1
    sheet[get_column_letter(int(i)) + "1"] = i - 1

for column in range(2, int(multiply_number) + 2):
    for row in range(2, int(multiply_number) + 2):
        sheet[get_column_letter(column) + str(row)] = f"=A{str(row)}*{get_column_letter(column)}1"

wb.save('multi_table.xlsx')